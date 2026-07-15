from __future__ import annotations

import argparse
import json
import math
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlparse

import numpy as np
from openpyxl import load_workbook
from PIL import Image, ImageFilter


ROOT = Path(__file__).resolve().parents[2]
APP_ROOT = Path(__file__).resolve().parents[1]
CONVERTED = ROOT / "tmp" / "converted"
IMAGE_ROOT = ROOT / "tmp" / "vision-images"
DOWNLOAD_CONFIG = ROOT / "tmp" / "vision-images.curl.conf"
DOWNLOAD_MAP = ROOT / "tmp" / "vision-images-map.json"
OUTPUT = APP_ROOT / "app" / "generated-data.json"


def norm(value: object) -> str:
    return " ".join(str(value or "").upper().split()).strip()


def canonical_plus(value: object) -> str:
    return "+".join(sorted(part.strip() for part in norm(value).split("+")))


def token_set(value: object) -> set[str]:
    ignored = {"100%", "100", "PERCENT", "THE"}
    return {part for part in re.findall(r"[A-Z0-9]+", norm(value)) if part not in ignored}


def jaccard(left: object, right: object) -> float:
    a, b = token_set(left), token_set(right)
    if not a and not b:
        return 1.0
    if not a or not b:
        return 0.0
    return len(a & b) / len(a | b)


def read_rows(path: Path, sheet: str) -> list[dict[str, object]]:
    ws = load_workbook(path, data_only=True, read_only=True)[sheet]
    raw = list(ws.iter_rows(values_only=True))
    headers = [str(value) for value in raw[0]]
    return [
        dict(zip(headers, row))
        for row in raw[1:]
        if any(value is not None for value in row)
    ]


def load_source_data():
    historical = read_rows(CONVERTED / "LAST TWO SEASON DATA_SAMPLE.xlsx", "Sheet1")
    upcoming = read_rows(CONVERTED / "upcomig season sample data.xlsx", "Sheet1")
    past_urls = read_rows(ROOT / "IMAGE URL FOR LAST SEASON - Project-1.xlsx", "Sheet2")
    upcoming_urls = read_rows(ROOT / "IMAGE URL FOR UPCOMING SEASON.xlsx", "Sheet1")
    return historical, upcoming, past_urls, upcoming_urls


def safe_basename(key: str, url: str) -> str:
    extension = Path(urlparse(url).path).suffix.lower() or ".jpg"
    return re.sub(r"[^A-Za-z0-9_.-]+", "_", key) + extension


def make_download_config() -> None:
    _, _, past_urls, upcoming_urls = load_source_data()
    entries: list[dict[str, str]] = []
    for group, rows, url_field in [
        ("past", past_urls, "DD"),
        ("upcoming", upcoming_urls, "link"),
    ]:
        for row in rows:
            key, url = str(row["CON"]), str(row[url_field])
            target = IMAGE_ROOT / group / safe_basename(key, url)
            entries.append({"group": group, "key": key, "url": url, "path": str(target)})

    IMAGE_ROOT.mkdir(parents=True, exist_ok=True)
    DOWNLOAD_CONFIG.parent.mkdir(parents=True, exist_ok=True)
    config_lines = [
        "silent",
        "show-error",
        "location",
        "fail",
        "retry = 2",
        "connect-timeout = 15",
        "max-time = 45",
        "create-dirs",
    ]
    for entry in entries:
        config_lines.extend(
            [
                f'url = "{entry["url"]}"',
                f'output = "{entry["path"]}"',
            ]
        )
    DOWNLOAD_CONFIG.write_text("\n".join(config_lines) + "\n", encoding="utf-8")
    DOWNLOAD_MAP.write_text(json.dumps(entries, indent=2), encoding="utf-8")
    print(f"Prepared {len(entries)} image downloads in {DOWNLOAD_CONFIG}")


def l2(values: np.ndarray) -> np.ndarray:
    magnitude = float(np.linalg.norm(values))
    return values / magnitude if magnitude else values


def image_feature(path: Path) -> list[float] | None:
    if not path.exists():
        return None
    try:
        image = Image.open(path).convert("RGB")
    except Exception:
        return None

    width, height = image.size
    crop = image.crop((int(width * 0.18), int(height * 0.10), int(width * 0.82), int(height * 0.68)))
    crop = crop.resize((48, 48), Image.Resampling.LANCZOS)
    rgb = np.asarray(crop, dtype=np.float32) / 255.0

    hist_parts = []
    for channel in range(3):
        hist, _ = np.histogram(rgb[:, :, channel], bins=12, range=(0.0, 1.0), density=False)
        hist_parts.append(hist.astype(np.float32))
    colour_hist = l2(np.concatenate(hist_parts)) * math.sqrt(0.34)

    grid = np.asarray(crop.resize((6, 6), Image.Resampling.BILINEAR), dtype=np.float32) / 255.0
    spatial = l2(grid.reshape(-1)) * math.sqrt(0.28)

    grey_image = crop.convert("L")
    grey = np.asarray(grey_image, dtype=np.float32) / 255.0
    grey = grey - float(grey.mean())
    spectrum = np.log1p(np.abs(np.fft.fft2(grey)))[:10, :10].astype(np.float32)
    texture = l2(spectrum.reshape(-1)) * math.sqrt(0.24)

    edge = np.asarray(grey_image.filter(ImageFilter.FIND_EDGES), dtype=np.float32) / 255.0
    edge_blocks = edge.reshape(6, 8, 6, 8).mean(axis=(1, 3))
    edges = l2(edge_blocks.reshape(-1)) * math.sqrt(0.14)

    return l2(np.concatenate([colour_hist, spatial, texture, edges])).round(6).tolist()


PATTERN_FAMILY = {
    "DIGITAL PRINT": "PRINT",
    "DISCHARGE PRINT": "PRINT",
    "PIGMENT PRINT": "PRINT",
    "PRINTS": "PRINT",
    "WBC DOBBY": "TEXTURE",
    "DOBBY/STRUCTURE": "TEXTURE",
}

COLOUR_FAMILY = {
    "NAVY BLUE": "BLUE",
    "SKY BLUE": "BLUE",
    "LIGHT BLUE": "BLUE",
    "PEACOCK BLUE": "BLUE",
    "TEAL": "BLUE",
    "INDIGO": "BLUE",
    "DARK GREEN": "GREEN",
    "LIGHT GREEN": "GREEN",
    "MINT": "GREEN",
    "CREAM": "NEUTRAL",
    "IVORY": "NEUTRAL",
    "BEIGE": "NEUTRAL",
    "KHAKI": "NEUTRAL",
    "STONE": "NEUTRAL",
    "WHITE": "NEUTRAL",
    "GREY": "GREY",
    "DARK GREY": "GREY",
    "LIGHT GREY": "GREY",
    "BLACK": "GREY",
    "MAROON": "RED",
    "CORAL": "RED",
    "RUST": "RED",
    "PINK": "RED",
    "OCHRE": "YELLOW",
    "LEMON": "YELLOW",
}


def categorical(left: object, right: object) -> float:
    return 1.0 if norm(left) == norm(right) else 0.0


def pattern_similarity(left: object, right: object) -> float:
    a, b = norm(left), norm(right)
    if a == b:
        return 1.0
    if PATTERN_FAMILY.get(a, a) == PATTERN_FAMILY.get(b, b):
        return 0.62
    return 0.08 if {a, b} <= {"CHECKS", "STRIPES", "PRINTS", "DIGITAL PRINT"} else 0.0


def colour_similarity(left: object, right: object) -> float:
    a, b = norm(left), norm(right)
    if a == b:
        return 1.0
    return 0.66 if COLOUR_FAMILY.get(a, a) == COLOUR_FAMILY.get(b, b) else 0.0


def attribute_score(upcoming: dict[str, object], historical: dict[str, object]):
    item_type = categorical(upcoming["SEGMENT1"], historical["Item Type"])
    values = {
        "category": item_type,
        "sleeve": categorical(upcoming["SEGMENT5"], historical["SLEEVS"]),
        "provision": categorical(upcoming["SEGMENT6"], historical["PROV"]),
        "pattern": pattern_similarity(upcoming["CAT1"], historical["CAT1"]),
        "range": 1.0 if canonical_plus(upcoming["CAT2"]) == canonical_plus(historical["CAT2"]) else 0.0,
        "fit": categorical(upcoming["CAT3"], historical["CAT3"]),
        "fabric": max(categorical(upcoming["CAT4"], historical["CAT4"]), jaccard(upcoming["CAT4"], historical["CAT4"])),
        "fashion": categorical(upcoming["CAT5"], historical["CAT5"]),
        "colour": colour_similarity(upcoming["COLOR"], historical["COLOR_NAME"]),
        "price": math.exp(-abs(math.log(max(float(upcoming["MRP"]), 1) / max(float(historical["MRP"]), 1))) / 0.30),
    }
    weights = {
        "category": 0.16,
        "sleeve": 0.07,
        "provision": 0.07,
        "pattern": 0.16,
        "range": 0.04,
        "fit": 0.14,
        "fabric": 0.14,
        "fashion": 0.02,
        "colour": 0.09,
        "price": 0.11,
    }
    score = sum(values[name] * weight for name, weight in weights.items())
    if item_type == 0:
        score *= 0.42
    return round(score, 4), {name: round(value, 3) for name, value in values.items()}


def cosine(left: list[float] | None, right: list[float] | None) -> float | None:
    if left is None or right is None:
        return None
    a, b = np.asarray(left, dtype=np.float32), np.asarray(right, dtype=np.float32)
    raw = float(np.dot(a, b) / max(float(np.linalg.norm(a) * np.linalg.norm(b)), 1e-9))
    # These compact visual descriptors share a high positive baseline. Calibrate
    # the cosine range so the UI score represents meaningful separation rather
    # than displaying every pair as an apparent 80-90% match.
    value = (raw - 0.62) / 0.37
    return round(max(0.0, min(1.0, value)), 4)


def clamp(value: float, lower: float, upper: float) -> float:
    return max(lower, min(upper, value))


def round_pack(value: float, pack: int = 25) -> int:
    return int(round(value / pack) * pack)


def recommendation(matches: list[dict[str, object]], history_by_id: dict[str, dict[str, object]]):
    top = matches[:5]
    numerator = 0.0
    denominator = 0.0
    for match in top:
        item = history_by_id[str(match["historicalId"])]
        base = (float(item["order"]) + float(item["dispatch"])) / 2
        factor = clamp(float(item["sellThrough"]) / 0.70, 0.65, 1.35)
        adjusted = base * factor
        weight = max(float(match["hybridScore"]), 0.01) ** 2
        numerator += adjusted * weight
        denominator += weight
    quantity = round_pack(numerator / denominator if denominator else 0)
    quantity = int(clamp(quantity, 100, 2000))
    top_score = float(top[0]["hybridScore"]) if top else 0.0
    mean_top_three = sum(float(item["hybridScore"]) for item in top[:3]) / max(len(top[:3]), 1)
    if top_score >= 0.82 and mean_top_three >= 0.72:
        confidence, spread = "High", 0.10
    elif top_score >= 0.67 and mean_top_three >= 0.58:
        confidence, spread = "Medium", 0.16
    else:
        confidence, spread = "Low", 0.25
    return {
        "quantity": quantity,
        "low": round_pack(quantity * (1 - spread)),
        "high": round_pack(quantity * (1 + spread)),
        "confidence": confidence,
    }


def build_data() -> None:
    historical_rows, upcoming_rows, past_urls, upcoming_urls = load_source_data()
    past_url_map = {str(row["CON"]): str(row["DD"]) for row in past_urls}
    upcoming_url_map = {str(row["CON"]): str(row["link"]) for row in upcoming_urls}
    download_entries = json.loads(DOWNLOAD_MAP.read_text(encoding="utf-8"))
    file_map = {(entry["group"], entry["key"]): Path(entry["path"]) for entry in download_entries}

    historical = []
    history_features: dict[str, list[float] | None] = {}
    for row in historical_rows:
        item_id = "-".join(norm(row[key]) for key in ["Item Type", "SORT", "COLOR", "SLEEVS"])
        url = past_url_map.get(item_id)
        feature = image_feature(file_map.get(("past", item_id), Path("__missing__")))
        history_features[item_id] = feature
        historical.append(
            {
                "id": item_id,
                "season": str(row["SEASON"]),
                "itemType": str(row["Item Type"]),
                "style": str(row["SORT"]),
                "colourCode": str(row["COLOR"]),
                "sleeve": str(row["SLEEVS"]),
                "provision": str(row["PROV"]),
                "pattern": str(row["CAT1"]),
                "range": str(row["CAT2"]),
                "fit": str(row["CAT3"]),
                "fabric": str(row["CAT4"]),
                "fashion": str(row["CAT5"]),
                "lifecycle": str(row["CAT6"]),
                "colour": str(row["COLOR_NAME"]),
                "mrp": int(row["MRP"]),
                "order": int(row["ORDER"]),
                "dispatch": int(row["DISPATCH"]),
                "sales": int(row["SALE"]),
                "sellThrough": round(float(row["SALE THRU"]), 4),
                "imageUrl": url,
                "hasVisualFeature": feature is not None,
            }
        )

    history_by_id = {item["id"]: item for item in historical}
    raw_history_by_id = {
        "-".join(norm(row[key]) for key in ["Item Type", "SORT", "COLOR", "SLEEVS"]): row
        for row in historical_rows
    }

    upcoming = []
    missing_images = []
    visual_scores = []
    attribute_scores = []
    for row in upcoming_rows:
        item_id = "-".join(norm(row[key]) for key in ["SEGMENT1", "SEGMENT2", "SEGMENT3"])
        url = upcoming_url_map.get(item_id)
        feature = image_feature(file_map.get(("upcoming", item_id), Path("__missing__")))
        if url is None:
            missing_images.append(item_id)

        matches = []
        for historical_item in historical:
            historical_id = historical_item["id"]
            attr, breakdown = attribute_score(row, raw_history_by_id[historical_id])
            visual = cosine(feature, history_features[historical_id])
            hybrid = attr if visual is None else attr * 0.65 + visual * 0.35
            matches.append(
                {
                    "historicalId": historical_id,
                    "attributeScore": attr,
                    "visualScore": visual,
                    "hybridScore": round(hybrid, 4),
                    "attributeBreakdown": breakdown,
                }
            )
            attribute_scores.append(attr)
            if visual is not None:
                visual_scores.append(visual)
        matches.sort(key=lambda item: float(item["hybridScore"]), reverse=True)

        upcoming.append(
            {
                "id": item_id,
                "itemType": str(row["SEGMENT1"]),
                "style": str(row["SEGMENT2"]),
                "colourCode": str(row["SEGMENT3"]),
                "sleeve": str(row["SEGMENT5"]),
                "provision": str(row["SEGMENT6"]),
                "pattern": str(row["CAT1"]),
                "range": str(row["CAT2"]),
                "fit": str(row["CAT3"]),
                "fabric": str(row["CAT4"]),
                "fashion": str(row["CAT5"]),
                "lifecycle": str(row["CAT6"]),
                "colour": str(row["COLOR"]),
                "mrp": int(row["MRP"]),
                "imageUrl": url,
                "hasVisualFeature": feature is not None,
                "matches": matches,
                "recommendation": recommendation(matches, history_by_id),
            }
        )

    confidence_counts = Counter(item["recommendation"]["confidence"] for item in upcoming)
    payload = {
        "meta": {
            "title": "Turtle Season Intelligence POC",
            "generatedAt": datetime.now(timezone.utc).isoformat(),
            "historicalItems": len(historical),
            "upcomingItems": len(upcoming),
            "historicalImageCoverage": sum(bool(item["imageUrl"]) for item in historical),
            "upcomingImageCoverage": sum(bool(item["imageUrl"]) for item in upcoming),
            "missingUpcomingImages": missing_images,
            "confidenceCounts": dict(confidence_counts),
            "attributeScoreRange": [round(min(attribute_scores), 3), round(max(attribute_scores), 3)],
            "visualScoreRange": [round(min(visual_scores), 3), round(max(visual_scores), 3)],
            "visualMethod": "Garment-region colour, structure, texture, and edge feature vector (POC)",
        },
        "historical": historical,
        "upcoming": upcoming,
    }
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")
    print(f"Wrote {OUTPUT} ({OUTPUT.stat().st_size / 1024:.1f} KB)")
    print(json.dumps(payload["meta"], indent=2))


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("mode", choices=["downloads", "build"])
    args = parser.parse_args()
    if args.mode == "downloads":
        make_download_config()
    else:
        build_data()
