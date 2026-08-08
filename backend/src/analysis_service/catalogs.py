"""Canonical spreadsheet ingestion and row-level validation for uploads."""

from __future__ import annotations

import csv
import shutil
import subprocess
import tempfile
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from PIL import Image, UnidentifiedImageError
from pyxlsb import open_workbook

from data_pipeline.feature_engineering import (
    build_historical_features,
    build_upcoming_features,
    exclude_upcoming_without_historical_item,
    exclude_zero_sales_rows,
)
from data_pipeline.preprocessing import (
    PreprocessingError,
    clean_product_identifier,
    normalize_text,
    parse_integer,
    parse_number,
    preprocess_rows,
    standardize_fabric,
)
from data_pipeline.schema import (
    CATEGORY_TYPE_MAP,
    SchemaStandardizationError,
    standardize_historical_schema,
    standardize_upcoming_schema,
)

# Upcoming workbooks are not required to carry a season column; this is the
# season stamped on rows that leave it blank.
DEFAULT_UPCOMING_SEASON = "SS27"

HISTORICAL_FIELDS = (
    "product_id",
    "season",
    "item_type",
    "style_code",
    "colour_code",
    "design",
    "category_type",
    "fabric",
    "colour",
    "total_order_quantity",
    "dispatch_quantity",
    "sales_quantity",
    "sell_through",
    "ageing_days",
    "weekly_sell_through",
)
UPCOMING_FIELDS = (
    "product_id",
    "season",
    "item_type",
    "style_code",
    "colour_code",
    "design",
    "category_type",
    "fabric",
    "colour",
    "collection_world",
)
SUPPORTED_IMAGE_SUFFIXES = {".jpg", ".jpeg", ".png", ".webp"}
SUPPORTED_CATALOGUE_SUFFIXES = {".csv", ".xlsx", ".xlsm", ".xls", ".xlsb", ".ods"}
OPENPYXL_SUFFIXES = {".xlsx", ".xlsm"}
CONVERTED_WORKBOOK_SUFFIXES = {".xls", ".ods"}
MAX_IMAGE_BYTES = 16 * 1024 * 1024
MAX_IMAGE_PIXELS = 40_000_000
MAX_CATALOGUE_ROWS = 100_000


@dataclass(frozen=True)
class ValidationIssue:
    catalog: str
    row: int | None
    product_id: str
    field: str
    code: str
    message: str
    severity: str = "error"
    record_type: str = "issue"

    def as_dict(self) -> dict[str, Any]:
        return {
            "recordType": self.record_type,
            "catalog": self.catalog,
            "row": self.row,
            "productId": self.product_id,
            "field": self.field,
            "code": self.code,
            "message": self.message,
            "severity": self.severity,
        }


@dataclass
class ValidatedCatalog:
    rows: list[dict[str, Any]]
    image_index: dict[str, Path]
    issues: list[ValidationIssue]
    report_records: list[ValidationIssue]


class CatalogValidationError(ValueError):
    def __init__(self, message: str, issues: list[ValidationIssue] | None = None):
        super().__init__(message)
        self.issues = issues or []


def _canonical_category(value: object) -> str:
    normalized = normalize_text(value)
    if normalized in set(CATEGORY_TYPE_MAP.values()):
        return normalized
    if normalized in CATEGORY_TYPE_MAP:
        return CATEGORY_TYPE_MAP[normalized]
    raise ValueError("must be FORMAL, CASUAL, DENIM, or CEREMONIAL")


def _validate_headers(actual: tuple[str, ...], catalog: str) -> tuple[str, ...]:
    expected = HISTORICAL_FIELDS if catalog == "historical" else UPCOMING_FIELDS
    missing = [field for field in expected if field not in actual]
    unexpected = [field for field in actual if field not in expected]
    duplicates = sorted({field for field in actual if actual.count(field) > 1})
    if missing or unexpected or duplicates:
        issue = ValidationIssue(
            catalog,
            1,
            "",
            "header",
            "schema_mismatch",
            f"missing={missing}; unexpected={unexpected}; duplicates={duplicates}",
        )
        raise CatalogValidationError("catalogue schema does not match the canonical template", [issue])
    return expected


def _read_csv(path: Path, catalog: str) -> list[dict[str, Any]]:
    try:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            actual = tuple(str(field or "").strip() for field in (reader.fieldnames or ()))
            expected = _validate_headers(actual, catalog)
            rows = []
            for index, row in enumerate(reader, start=1):
                if index > MAX_CATALOGUE_ROWS:
                    raise CatalogValidationError(f"catalogue exceeds the {MAX_CATALOGUE_ROWS} row limit")
                normalized = {str(key or "").strip(): value for key, value in row.items()}
                if any(value not in (None, "") for value in normalized.values()):
                    rows.append({field: normalized.get(field) for field in expected})
            return rows
    except UnicodeDecodeError as exc:
        raise CatalogValidationError("CSV must use UTF-8 encoding") from exc


def _read_xlsx(path: Path, catalog: str) -> list[dict[str, Any]]:
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise CatalogValidationError("workbook is corrupt, encrypted, or unsupported") from exc
    try:
        worksheet = workbook["Sheet1"] if "Sheet1" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        rows = worksheet.iter_rows(values_only=True)
        try:
            raw_headers = next(rows)
        except StopIteration as exc:
            raise CatalogValidationError("workbook contains no rows") from exc
        values_rows: list[tuple[Any, ...]] = []
        for index, values in enumerate(rows, start=1):
            if index > MAX_CATALOGUE_ROWS:
                raise CatalogValidationError(f"catalogue exceeds the {MAX_CATALOGUE_ROWS} row limit")
            if any(value not in (None, "") for value in values):
                values_rows.append(tuple(values))
        return _rows_from_workbook_values(tuple(raw_headers), values_rows, catalog)
    finally:
        workbook.close()


def _read_xlsb(path: Path, catalog: str) -> list[dict[str, Any]]:
    """Read binary Excel workbooks directly, without spawning LibreOffice."""

    try:
        with open_workbook(str(path)) as workbook:
            sheet_name = "Sheet1" if "Sheet1" in workbook.sheets else workbook.sheets[0]
            with workbook.get_sheet(sheet_name) as worksheet:
                rows = worksheet.rows()
                try:
                    raw_headers = tuple(cell.v for cell in next(rows))
                except StopIteration as exc:
                    raise CatalogValidationError("workbook contains no rows") from exc
                values_rows: list[tuple[Any, ...]] = []
                for index, row in enumerate(rows, start=1):
                    if index > MAX_CATALOGUE_ROWS:
                        raise CatalogValidationError(f"catalogue exceeds the {MAX_CATALOGUE_ROWS} row limit")
                    values = tuple(cell.v for cell in row)
                    if any(value not in (None, "") for value in values):
                        values_rows.append(values)
    except CatalogValidationError:
        raise
    except Exception as exc:
        raise CatalogValidationError("workbook is corrupt, encrypted, or unsupported") from exc
    return _rows_from_workbook_values(raw_headers, values_rows, catalog)


def _rows_from_workbook_values(
    raw_headers: tuple[Any, ...],
    values_rows: list[tuple[Any, ...]],
    catalog: str,
) -> list[dict[str, Any]]:
    actual = tuple(str(value or "").strip() for value in raw_headers)
    try:
        expected = _validate_headers(actual, catalog)
    except CatalogValidationError as canonical_error:
        try:
            return _standardize_legacy_workbook(raw_headers, values_rows, catalog)
        except (KeyError, PreprocessingError, SchemaStandardizationError, ValueError) as exc:
            raise canonical_error from exc
    return [
        {field: dict(zip(actual, values, strict=True)).get(field) for field in expected}
        for values in values_rows
    ]


def _standardize_legacy_workbook(
    raw_headers: tuple[Any, ...],
    values_rows: list[tuple[Any, ...]],
    catalog: str,
) -> list[dict[str, Any]]:
    """Accept Turtle's established XLS/XLSB layouts alongside canonical templates."""

    seen: Counter[str] = Counter()
    headers: list[str] = []
    for value in raw_headers:
        header = normalize_text(value)
        seen[header] += 1
        headers.append(header if seen[header] == 1 else f"{header}__{seen[header]}")
    source_rows = [dict(zip(headers, values, strict=True)) for values in values_rows]
    if catalog == "historical":
        if source_rows and "IMAGE ID" in source_rows[0]:
            renamed_columns = {
                "IMAGE ID": "CON",
                "SEGMENT1": "ITEM TYPE",
                "SEGMENT2": "SORT",
                "SEGMENT3": "COLOR",
                "COLOR_NAME": "COLOR__2",
                "SELL THROUGH": "SELL THR",
                "WEEKLY SELL THROUGH": "WKLY SELL THRU",
            }
            normalized_rows: list[dict[str, Any]] = []
            for row in source_rows:
                normalized = dict(row)
                for source, target in renamed_columns.items():
                    normalized[target] = normalized.pop(source)
                normalized_rows.append(normalized)
            source_rows = normalized_rows
        cleaned, _report = preprocess_rows("Historical", source_rows, "CON")
        standardized = standardize_historical_schema(cleaned)
        return [{field: row.get(field) for field in HISTORICAL_FIELDS} for row in standardized]

    if source_rows and "IMAGE ID" in source_rows[0]:
        normalized_rows = []
        for row in source_rows:
            normalized = dict(row)
            normalized["CC (SEG-1+2+3)"] = normalized.pop("IMAGE ID")
            normalized["COLOUR NAME"] = normalized.pop("COLOR_NAME")
            normalized["CAT-3"] = normalized.pop("CAT3")
            normalized_rows.append(normalized)
        source_rows = normalized_rows
    cleaned, _report = preprocess_rows("Upcoming", source_rows, "CC (SEG-1+2+3)")
    standardized = standardize_upcoming_schema(cleaned, DEFAULT_UPCOMING_SEASON)
    return [{field: row.get(field) for field in UPCOMING_FIELDS} for row in standardized]


def _read_converted_workbook(path: Path, catalog: str) -> list[dict[str, Any]]:
    executable = shutil.which("soffice")
    if executable is None:
        raise CatalogValidationError(f"LibreOffice is required to read {path.suffix.lower()} catalogue files")
    with tempfile.TemporaryDirectory(prefix="turtle-workbook-", dir=path.parent) as temporary_name:
        temporary = Path(temporary_name)
        profile = temporary / "libreoffice-profile"
        command = [
            executable,
            f"-env:UserInstallation={profile.resolve().as_uri()}",
            "--headless",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(temporary),
            str(path),
        ]
        try:
            completed = subprocess.run(command, check=True, capture_output=True, text=True, timeout=120)
        except (subprocess.CalledProcessError, subprocess.TimeoutExpired) as exc:
            raise CatalogValidationError(f"failed to convert {path.suffix.lower()} workbook") from exc
        candidates = sorted(temporary.glob("*.xlsx"))
        if len(candidates) != 1:
            detail = completed.stderr.strip() or completed.stdout.strip()
            raise CatalogValidationError(f"LibreOffice did not produce a readable workbook: {detail}")
        return _read_xlsx(candidates[0], catalog)


def _read_catalogue(path: Path, catalog: str) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix not in SUPPORTED_CATALOGUE_SUFFIXES:
        raise CatalogValidationError(
            f"unsupported catalogue format {suffix or '<none>'}; use CSV, XLSX, XLSM, XLS, XLSB, or ODS"
        )
    if suffix == ".csv":
        return _read_csv(path, catalog)
    if suffix in OPENPYXL_SUFFIXES:
        return _read_xlsx(path, catalog)
    if suffix == ".xlsb":
        return _read_xlsb(path, catalog)
    if suffix in CONVERTED_WORKBOOK_SUFFIXES:
        return _read_converted_workbook(path, catalog)
    raise CatalogValidationError(f"unsupported catalogue format: {suffix}")


def _validate_image_directory(directory: Path, catalog: str) -> tuple[dict[str, Path], list[ValidationIssue]]:
    index: dict[str, Path] = {}
    issues: list[ValidationIssue] = []
    if not directory.is_dir():
        return index, issues
    for path in sorted(directory.rglob("*")):
        if not path.is_file():
            continue
        suffix = path.suffix.lower()
        product_id = path.stem.upper()
        if suffix not in SUPPORTED_IMAGE_SUFFIXES:
            issues.append(ValidationIssue(catalog, None, product_id, "image", "unsupported_image", path.name))
            continue
        if product_id in index:
            issues.append(
                ValidationIssue(catalog, None, product_id, "image", "duplicate_image", "duplicate image filename stem")
            )
            continue
        try:
            if path.stat().st_size > MAX_IMAGE_BYTES:
                raise ValueError(f"image exceeds {MAX_IMAGE_BYTES} byte limit")
            with Image.open(path) as image:
                image.verify()
                if image.width * image.height > MAX_IMAGE_PIXELS:
                    raise ValueError(f"image exceeds {MAX_IMAGE_PIXELS} pixel limit")
                if (image.format or "").upper() not in {"JPEG", "PNG", "WEBP"}:
                    raise ValueError("decoded image format is unsupported")
        except (OSError, ValueError, UnidentifiedImageError) as exc:
            issues.append(ValidationIssue(catalog, None, product_id, "image", "invalid_image", str(exc)))
            continue
        index[product_id] = path
    return index, issues


def validate_catalog(catalogue_path: Path, image_directory: Path, catalog: str) -> ValidatedCatalog:
    raw_rows = _read_catalogue(catalogue_path, catalog)
    issues: list[ValidationIssue] = []
    image_index, image_issues = _validate_image_directory(image_directory, catalog)
    issues.extend(image_issues)
    cleaned: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row_number, raw in enumerate(raw_rows, start=2):
        source_id = str(raw.get("product_id") or "").strip()
        row_issues: list[ValidationIssue] = []
        try:
            product_id = clean_product_identifier(source_id)
        except PreprocessingError as exc:
            row_issues.append(
                ValidationIssue(catalog, row_number, source_id, "product_id", "invalid_identifier", str(exc))
            )
            product_id = source_id
        if product_id in seen:
            row_issues.append(
                ValidationIssue(
                    catalog, row_number, product_id, "product_id", "duplicate_identifier", "duplicate product_id"
                )
            )
        try:
            category = _canonical_category(raw.get("category_type"))
        except ValueError as exc:
            row_issues.append(
                ValidationIssue(catalog, row_number, product_id, "category_type", "invalid_category", str(exc))
            )
            category = ""
        numeric: dict[str, int | float] = {}
        if catalog == "historical":
            for field in ("total_order_quantity", "dispatch_quantity", "sales_quantity", "ageing_days"):
                try:
                    numeric[field] = parse_integer(raw.get(field))
                    if numeric[field] < 0:
                        raise ValueError("must not be negative")
                except (TypeError, ValueError):
                    row_issues.append(
                        ValidationIssue(
                            catalog, row_number, product_id, field, "invalid_number", "must be a non-negative number"
                        )
                    )
            for field in ("sell_through", "weekly_sell_through"):
                try:
                    numeric[field] = parse_number(raw.get(field))
                    if numeric[field] < 0:
                        raise ValueError("must not be negative")
                except (TypeError, ValueError):
                    row_issues.append(
                        ValidationIssue(
                            catalog, row_number, product_id, field, "invalid_number", "must be a non-negative number"
                        )
                    )
        if row_issues:
            issues.extend(row_issues)
            continue
        seen.add(product_id)
        cleaned.append(
            {
                **raw,
                **numeric,
                "product_id": product_id,
                "season": normalize_text(raw.get("season")),
                "item_type": normalize_text(raw.get("item_type")),
                "style_code": normalize_text(raw.get("style_code")),
                "colour_code": normalize_text(raw.get("colour_code")),
                "design": normalize_text(raw.get("design")),
                "category_type": category,
                "fabric": standardize_fabric(raw.get("fabric")),
                "colour": normalize_text(raw.get("colour")),
                **({"collection_world": normalize_text(raw.get("collection_world"))} if catalog == "upcoming" else {}),
            }
        )
    if not cleaned:
        raise CatalogValidationError(f"{catalog} catalogue contains no valid rows", issues)
    valid_ids = {str(row["product_id"]).upper() for row in cleaned}
    matched_ids = valid_ids & set(image_index)
    missing_ids = valid_ids - set(image_index)
    unused_image_ids = set(image_index) - valid_ids
    matched_records = [
        ValidationIssue(
            catalog,
            None,
            product_id,
            "image",
            "matched_image",
            f"matched uploaded file {image_index[product_id].name}",
            "passed",
            "check",
        )
        for product_id in sorted(matched_ids)
    ]
    for product_id in sorted(missing_ids):
        issues.append(
            ValidationIssue(
                catalog,
                None,
                product_id,
                "image",
                "missing_image",
                "no supported image filename matches product_id",
                "warning",
            )
        )
    for product_id in sorted(unused_image_ids):
        issues.append(
            ValidationIssue(
                catalog,
                None,
                product_id,
                "image",
                "unused_image",
                f"uploaded file {image_index[product_id].name} does not match any catalogue product_id",
                "warning",
            )
        )
    summary = ValidationIssue(
        catalog,
        None,
        "",
        "image",
        "image_coverage_summary",
        (
            f"{len(matched_ids)} of {len(valid_ids)} catalogue products matched an uploaded image; "
            f"{len(missing_ids)} missing; {len(unused_image_ids)} uploaded image files unused. "
            "The rows below include every matched product plus each validation issue."
        ),
        "info",
        "summary",
    )
    return ValidatedCatalog(cleaned, image_index, issues, [summary, *matched_records, *issues])


def build_feature_source(
    historical_rows: list[dict[str, Any]],
    upcoming_rows: list[dict[str, Any]],
    *,
    historical_image_ids: set[str],
    upcoming_image_ids: set[str],
    historical_version_id: str,
    build_id: str,
) -> dict[str, Any]:
    historical_rows, zero_sales = exclude_zero_sales_rows(historical_rows)
    upcoming_rows, unseen = exclude_upcoming_without_historical_item(historical_rows, upcoming_rows)
    history, duplicates = build_historical_features(historical_rows)
    upcoming, unseen_features = build_upcoming_features(
        upcoming_rows,
        {normalize_text(row["item_type"]) for row in historical_rows},
    )
    for item in history:
        source_id = str(item["sourceId"])
        item["imageUrl"] = (
            f"/api/builds/{build_id}/images/historical/{source_id}"
            if source_id.upper() in historical_image_ids
            else None
        )
        item["hasVisualFeature"] = False
    for item in upcoming:
        product_id = str(item["id"])
        item["imageUrl"] = (
            f"/api/builds/{build_id}/images/upcoming/{product_id}" if product_id.upper() in upcoming_image_ids else None
        )
        item["hasVisualFeature"] = False
    return {
        "meta": {
            "title": "Turtle Season Intelligence AI",
            "dataMode": "uploaded",
            "historicalVersionId": historical_version_id,
            "upcomingSeason": normalize_text(upcoming_rows[0].get("season")),
            "historicalItems": len(history),
            "upcomingItems": len(upcoming),
            "historicalImageCoverage": sum(bool(item["imageUrl"]) for item in history),
            "upcomingImageCoverage": sum(bool(item["imageUrl"]) for item in upcoming),
            "missingUpcomingImages": [item["id"] for item in upcoming if not item["imageUrl"]],
            "dataQuality": {
                "duplicateHistoricalRowsRemoved": duplicates,
                "zeroSalesHistoricalRowsExcluded": zero_sales,
                "upcomingRowsExcludedUnseenItem": unseen,
                "upcomingWithoutHistoricalItem": unseen_features,
            },
        },
        "historical": history,
        "upcoming": upcoming,
    }
