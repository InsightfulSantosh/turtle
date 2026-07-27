"""Workbook discovery, conversion and row ingestion."""

from __future__ import annotations

import shutil
import subprocess
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from data_pipeline.settings import PipelineSettings


@dataclass(frozen=True)
class IngestedData:
    historical_rows: list[dict[str, Any]]
    upcoming_rows: list[dict[str, Any]]


class WorkbookIngestor:
    """Reads the supplied XLSB workbooks through a deterministic XLSX cache."""

    def __init__(self, settings: PipelineSettings):
        self.settings = settings

    def converted_path(self, source: Path) -> Path:
        return self.settings.converted_root / f"{source.stem}.xlsx"

    def ingest(self) -> IngestedData:
        sources = [
            self.settings.historical_source,
            self.settings.upcoming_source,
        ]
        self._validate_sources(sources)
        self._convert_stale_workbooks(sources)
        return IngestedData(
            historical_rows=self._read_rows(
                self.converted_path(self.settings.historical_source)
            ),
            upcoming_rows=self._read_rows(
                self.converted_path(self.settings.upcoming_source)
            ),
        )

    @staticmethod
    def _validate_sources(sources: list[Path]) -> None:
        missing = [source for source in sources if not source.is_file()]
        if missing:
            joined = ", ".join(str(source) for source in missing)
            raise FileNotFoundError(f"Missing source workbook(s): {joined}")

    def _convert_stale_workbooks(self, sources: list[Path]) -> None:
        stale = [
            source
            for source in sources
            if not self.converted_path(source).exists()
            or self.converted_path(source).stat().st_mtime < source.stat().st_mtime
        ]
        if not stale:
            return

        executable = shutil.which("soffice")
        if not executable:
            raise RuntimeError(
                "LibreOffice is required to convert the supplied XLSB workbooks"
            )

        self.settings.converted_root.mkdir(parents=True, exist_ok=True)
        profile = self.settings.libreoffice_profile.resolve()
        command = [
            executable,
            f"-env:UserInstallation={profile.as_uri()}",
            "--headless",
            "--convert-to",
            "xlsx",
            "--outdir",
            str(self.settings.converted_root),
            *(str(source) for source in stale),
        ]
        subprocess.run(command, check=True)

    def _read_rows(self, path: Path) -> list[dict[str, Any]]:
        worksheet = load_workbook(
            path,
            read_only=True,
            data_only=True,
        )[self.settings.sheet_name]
        raw_rows = worksheet.iter_rows(values_only=True)
        raw_headers = [self._header(value) for value in next(raw_rows)]
        seen: Counter[str] = Counter()
        headers: list[str] = []
        for header in raw_headers:
            seen[header] += 1
            headers.append(
                header if seen[header] == 1 else f"{header}__{seen[header]}"
            )
        return [
            dict(zip(headers, row, strict=True))
            for row in raw_rows
            if any(value not in (None, "") for value in row)
        ]

    @staticmethod
    def _header(value: object) -> str:
        return " ".join(str(value or "").upper().split()).strip()
